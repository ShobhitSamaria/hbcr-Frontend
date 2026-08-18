#!/usr/bin/env python3
"""
extract-icd10-workbook.py — one-time extraction of the project's ICD-10
workbook (Docs/ICD10_Cancer_Topography_Morphology.xlsx) into the JSON file that
Backend/scripts/seed-icd10.ts imports.

Only content that actually exists in the workbook is emitted:
  - code RANGES + category names (Code Ranges Reference sheet, plus the
    C00-C96 malignant behaviour block from Rule A),
  - RULES A-G (the instruction-manual rule text),
  - worked EXAMPLES (scenario + printed resulting-code text),
  - CODE MENTIONS (every individual code named by an example, with the
    example as context — the workbook contains no code-to-description
    mappings, so individual codes are never paired with invented terms).

Usage:
  python3 scripts/extract-icd10-workbook.py [path/to/workbook.xlsx]

Output: scripts/output/icd10-reference.json
Requires: openpyxl (already used for the earlier workbook inspection).
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

DEFAULT_XLSX = (
    "/Users/shobhitsamaria/Desktop/DOITC Code/HBCR/Docs/"
    "ICD10_Cancer_Topography_Morphology.xlsx"
)
OUT_PATH = Path(__file__).resolve().parent / "output" / "icd10-reference.json"

# Individual ICD-10 codes as printed in the workbook's "Resulting Code(s)"
# cells (e.g. "C18.9", "D48.6", "C80.0", "C19", "C23").
CODE_RE = re.compile(r"\b([CD]\d{2}(?:\.\d+)?)\b")
# Code ranges as printed, e.g. "C00–C75" / "C71" (en-dash separator).
RANGE_RE = re.compile(r"^([CD]\d{2})(?:–([CD]\d{2}))?$")


def cells(row: tuple) -> list[str]:
    """Non-empty, stripped cell values of a workbook row."""
    return [str(v).strip() for v in row if v is not None and str(v).strip()]


def first_row(ws, predicate, start=1):
    """Return (index, cells) of the first row (1-based) matching predicate."""
    for idx, row in enumerate(ws.iter_rows(min_row=start, values_only=True), start=start):
        c = cells(row)
        if predicate(c):
            return idx, c
    return None, None


def collect_ranges(ws):
    """Code Ranges Reference sheet -> [{code, title}]."""
    out = []
    for row in ws.iter_rows(values_only=True):
        c = cells(row)
        if len(c) >= 2 and RANGE_RE.match(c[0]) and c[1] != "Category":
            out.append({"code": c[0], "title": c[1]})
    return out


def collect_behaviour_ranges(ws):
    """Rule A behaviour blocks -> extra ranges not already in the reference.
    The title combines the behaviour label and its meaning verbatim, e.g.
    "Malignant — Invades surrounding tissue or spreads (metastasizes)..."""
    blocks = []
    for row in ws.iter_rows(values_only=True):
        c = cells(row)
        if len(c) >= 3 and c[0] != "Behaviour" and RANGE_RE.match(c[2]):
            blocks.append({"code": c[2], "title": f"{c[0]} — {c[1]}"})
    return blocks


def collect_rule_a(ws):
    """Numbered 'How to Determine Behaviour' instructions (Rule A)."""
    start, _ = first_row(ws, lambda c: c and c[0] == "How to Determine Behaviour")
    if not start:
        return ""
    parts = []
    for row in ws.iter_rows(min_row=start + 1, values_only=True):
        c = cells(row)
        if c and re.match(r"^\d+\.\s", c[0]):
            parts.append(c[0])
    return " ".join(parts)


def collect_rule_b(ws):
    """Common-sites-of-metastases list + decision sub-rules (Rule B)."""
    parts = []

    start, _ = first_row(ws, lambda c: c and "Common Sites of Metastases" in c[0])
    if start:
        end, _ = first_row(ws, lambda c: c and c[0] == "Decision Rules", start=start + 1)
        sites = []
        for row in ws.iter_rows(min_row=start + 1, max_row=(end - 1 if end else None), values_only=True):
            c = cells(row)
            if c:
                sites.append(c[0])
        if sites:
            parts.append("Common sites of metastases: " + ", ".join(sites) + ".")

    start, _ = first_row(ws, lambda c: c and c[0] == "Decision Rules")
    if start:
        for row in ws.iter_rows(min_row=start + 1, values_only=True):
            c = cells(row)
            if len(c) >= 2 and c[0] != "Sub-rule":
                parts.append(f"{c[0]} — {c[1]}")
    return " ".join(parts)


def collect_rule_cf(ws):
    """Rules C-F rows (rule label — text)."""
    start, _ = first_row(ws, lambda c: len(c) >= 2 and c[1] == "Text")
    if not start:
        return ""
    parts = []
    for row in ws.iter_rows(min_row=start + 1, values_only=True):
        c = cells(row)
        if len(c) >= 2:
            parts.append(f"{c[0]} — {c[1]}")
    return " ".join(parts)


def collect_rule_g(ws):
    """Intro paragraph + phrasing sub-rules (Rule G)."""
    parts = []
    start, _ = first_row(ws, lambda c: c and c[0] == "Phrasing")
    if start:
        intro_row = next(
            (cells(r) for r in ws.iter_rows(min_row=2, max_row=start - 1, values_only=True) if cells(r)),
            None,
        )
        if intro_row:
            parts.append(intro_row[0])
        for row in ws.iter_rows(min_row=start + 1, values_only=True):
            c = cells(row)
            if len(c) >= 2:
                parts.append(f"{c[0]} — {c[1]}")
    return " ".join(parts)


def collect_examples(ws):
    """All Examples sheet -> [{exampleNo, rule, scenario, resultText, codes}]."""
    out = []
    for row in ws.iter_rows(values_only=True):
        c = cells(row)
        if len(c) >= 4 and c[0] != "Example #":
            example_no, rule, scenario, result_text = c[0], c[1], c[2], c[3]
            codes = sorted({m for m in CODE_RE.findall(result_text)})
            out.append(
                {
                    "exampleNo": example_no,
                    "rule": rule,
                    "scenario": scenario,
                    "resultText": result_text,
                    "codes": codes,
                }
            )
    return out


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_XLSX)
    if not xlsx.is_file():
        print(f"[extract-icd10] error: workbook not found at {xlsx}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    ranges = collect_ranges(wb["Code Ranges Reference"])
    known = {r["code"] for r in ranges}
    for block in collect_behaviour_ranges(wb["Behaviour (Rule A)"]):
        if block["code"] not in known:
            ranges.append(block)
            known.add(block["code"])

    rules = [
        {
            "ruleId": "A",
            "title": "Behaviour: Malignant, In Situ, Benign, or Uncertain/Unknown",
            "text": collect_rule_a(wb["Behaviour (Rule A)"]),
            "source": "Behaviour (Rule A)",
        },
        {
            "ruleId": "B",
            "title": "Malignant Neoplasms: Primary or Secondary?",
            "text": collect_rule_b(wb["Primary vs Secondary (Rule B)"]),
            "source": "Primary vs Secondary (Rule B)",
        },
        {
            "ruleId": "C-F",
            "title": "Multiple Primaries, Vague Sites, Unknown Primary, Overlapping Sites",
            "text": collect_rule_cf(wb["Other Rules (C-F)"]),
            "source": "Other Rules (C-F)",
        },
        {
            "ruleId": "G",
            "title": "Parsing the word 'Metastatic'",
            "text": collect_rule_g(wb["Metastatic Terminology (Rule G)"]),
            "source": "Metastatic Terminology (Rule G)",
        },
    ]

    examples = collect_examples(wb["All Examples"])

    mentions = []
    for i, ex in enumerate(examples):
        for code in ex["codes"]:
            mentions.append(
                {
                    "code": code,
                    "exampleNo": ex["exampleNo"],
                    "rule": ex["rule"],
                    "scenario": ex["scenario"],
                    "sortOrder": i,
                }
            )

    doc = {
        "meta": {
            "source": str(xlsx),
            "sheetNames": wb.sheetnames,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "note": (
                "Extracted verbatim from the ICD-10 workbook (ICD-10 Volume 2 "
                "instruction-manual content). Contains code ranges + category "
                "names, rules and worked examples only. Individual codes are "
                "NOT mapped to descriptions — the workbook has none."
            ),
        },
        "ranges": ranges,
        "rules": rules,
        "examples": examples,
        "codeMentions": mentions,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    empty = [r for r in rules if not r["text"]]
    print(f"[extract-icd10] wrote {OUT_PATH}")
    print(f"[extract-icd10] ranges={len(ranges)} rules={len(rules)} "
          f"examples={len(examples)} codeMentions={len(mentions)}")
    if empty:
        print(f"[extract-icd10] WARNING: empty rule text for {[r['ruleId'] for r in empty]}",
              file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
